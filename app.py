from flask import Flask, render_template, request, send_file, session
import pandas as pd
from rapidfuzz import fuzz
import os
import pickle
import uuid
import re
import openpyxl
from difflib import SequenceMatcher
from fuzzywuzzy import fuzz, process
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell


app = Flask(__name__)
app.secret_key = 'supersecretkey'

def fuzzy_match(a, b):
    return SequenceMatcher(None, a, b).ratio()


def load_sheet_with_header_offset(xls, sheet_name, header_row=8):
    # baca sheet sebagai list tuples
    data = list(xls.parse(sheet_name, header=None).values)
    # header ada di baris ke-(header_row + 1), indexing 0-based
    columns = data[header_row]
    rows = data[header_row+1:]
    df = pd.DataFrame(rows, columns=columns)
    df.columns = df.columns.str.strip()  # bersihkan spasi kolom
    

    return df


# ==================== FITUR 1: BANDINGKAN EXCEL (OPTIMASI CEPAT) ====================
def compare_sheets_fast(df1, df2):
    df_result = df1.copy()
    df_result['Status Cocok'] = 'Tidak ditemukan'
    df_result['Metode Cocok'] = ''

    # Pastikan tipe kolom string dan strip spasi
    for col in ['UUID', 'ISBN Cetak', 'ISBN Elektronik*', 'Judul*', 'Anak Judul']:
        if col in df1.columns:
            df1[col] = df1[col].astype(str).str.strip()
        if col in df2.columns:
            df2[col] = df2[col].astype(str).str.strip()

    df2_uuid = df2.set_index('UUID', drop=False) if 'UUID' in df2.columns else pd.DataFrame()
    df2_isbn = df2.set_index('ISBN Cetak', drop=False) if 'ISBN Cetak' in df2.columns else pd.DataFrame()
    df2_eisbn = df2.set_index('ISBN Elektronik*', drop=False) if 'ISBN Elektronik*' in df2.columns else pd.DataFrame()

    judul_list = df2['Judul*'].dropna().astype(str).tolist() if 'Judul*' in df2.columns else []
    anak_judul_list = df2['Anak Judul'].dropna().astype(str).tolist() if 'Anak Judul' in df2.columns else []

    for idx, row1 in df_result.iterrows():
        matched_row = None
        match_method = None

        uuid = row1.get('UUID')
        isbn = row1.get('ISBN Cetak')
        eisbn = row1.get('ISBN Elektronik*')
        judul = row1.get('Judul*')

        print(f"Periksa baris {idx}: UUID={uuid}, ISBN={isbn}, EISBN={eisbn}, Judul={judul}")

        # 1️⃣ Cek UUID
        if pd.notna(uuid) and uuid in df2_uuid.index:
            matched_row = df2_uuid.loc[uuid]
            if isinstance(matched_row, pd.DataFrame):
                matched_row = matched_row.iloc[0]
            match_method = 'UUID'

        # 2️⃣ Cek ISBN Cetak
        elif pd.notna(isbn) and isbn in df2_isbn.index:
            matched_row = df2_isbn.loc[isbn]
            if isinstance(matched_row, pd.DataFrame):
                matched_row = matched_row.iloc[0]
            match_method = 'ISBN Cetak'

        # 3️⃣ Cek ISBN Elektronik
        elif pd.notna(eisbn) and eisbn in df2_eisbn.index:
            matched_row = df2_eisbn.loc[eisbn]
            if isinstance(matched_row, pd.DataFrame):
                matched_row = matched_row.iloc[0]
            match_method = 'ISBN Elektronik*'

        # 4️⃣ Fuzzy match Judul* (dengan validasi Anak Judul sama)
        elif pd.notna(judul) and judul_list:
            possible_matches = None
            best_match = process.extractOne(str(judul), judul_list, scorer=fuzz.ratio)
            if best_match:
                print(f"Judul fuzzy match: {best_match[0]} dengan skor {best_match[1]}")
            if best_match and best_match[1] > 70:
                possible_matches = df2[df2['Judul*'] == best_match[0]]

                if 'Anak Judul' in df2.columns and 'Anak Judul' in df1.columns:
                    anak1 = str(row1.get('Anak Judul', '')).strip()
                    for _, candidate in possible_matches.iterrows():
                        anak2 = str(candidate.get('Anak Judul', '')).strip()
                        if anak1 == anak2:  # sama persis
                            matched_row = candidate
                            match_method = 'Judul* + Anak Judul cocok'
                            break
                    if matched_row is None:
                        print(f"Judul cocok tapi Anak Judul beda → dianggap tidak cocok")
                else:
                    # Kalau tidak ada kolom Anak Judul di salah satu file
                    matched_row = possible_matches.iloc[0]
                    match_method = 'Judul* (Fuzzy)'

        # 5️⃣ Fuzzy match Anak Judul (kalau belum ketemu)
        if matched_row is None and pd.notna(judul) and anak_judul_list:
            best_match_anak = process.extractOne(str(judul), anak_judul_list, scorer=fuzz.ratio)
            if best_match_anak:
                print(f"Anak Judul fuzzy match: {best_match_anak[0]} dengan skor {best_match_anak[1]}")
            if best_match_anak and best_match_anak[1] > 70:
                matched_row = df2[df2['Anak Judul'] == best_match_anak[0]].iloc[0]
                match_method = 'Anak Judul (Fuzzy)'

        # Simpan hasil
        if matched_row is not None:
            df_result.at[idx, 'Status Cocok'] = 'Ditemukan'
            df_result.at[idx, 'Metode Cocok'] = match_method
            print(f"Baris {idx} ditemukan dengan metode {match_method}")

    return df_result
# ==================== PANGGIL FUNGSI BANDINGKAN ====================
def compare_excels(file_a, file_b, mode, sheet_a_name=None, sheet_b_name=None):
    import re
    import pandas as pd

    xls_a = pd.ExcelFile(file_a)
    xls_b = pd.ExcelFile(file_b)
    results = {}

    skip_sheets = ['Hasil Seleksi', 'Referensi', 'Form Pengadaan']

    def extract_sheet_name(sheet):
        return ''.join(re.findall(r'[A-Za-z ]+', sheet)).strip()


    print("[DEBUG] Mulai compare_excels")
    print(f"[DEBUG] File A sheets: {xls_a.sheet_names}")
    print(f"[DEBUG] File B sheets: {xls_b.sheet_names}")
    print(f"[DEBUG] Mode: {mode}")

    if mode == "single":
        if not sheet_a_name or not sheet_b_name:
            raise ValueError("Nama Sheet A dan B wajib diisi untuk mode 'single'.")
        df_a = load_sheet_with_header_offset(xls_a, sheet_a_name, header_row=8)
        df_b = load_sheet_with_header_offset(xls_b, sheet_b_name, header_row=8)
        key = f"{extract_sheet_name(sheet_a_name)} vs {extract_sheet_name(sheet_b_name)}"
        print(f"[DEBUG] Membandingkan sheet single: {key}")
        results[key] = compare_sheets_fast(df_a, df_b)

    elif mode == "multi":
        if not sheet_a_name:
            raise ValueError("Nama Sheet A wajib diisi untuk mode 'multi'.")
        df_a = load_sheet_with_header_offset(xls_a, sheet_a_name, header_row=8)
        for sheet in xls_b.sheet_names:
            if sheet in skip_sheets:
                print(f"[DEBUG] Lewat sheet {sheet} karena skip_sheets")
                continue
            print(f"[DEBUG] Membandingkan sheet multi: {sheet_a_name} vs {sheet}")
            df_b = load_sheet_with_header_offset(xls_b, sheet, header_row=8)
            key = f"{extract_sheet_name(sheet_a_name)} vs {extract_sheet_name(sheet)}"
            results[key] = compare_sheets_fast(df_a, df_b)

    elif mode == "multi-matching":
        print("[DEBUG] Mode: multi-matching berdasarkan nama sheet tanpa angka prefix")

        # Mapping clean_name -> original sheet name untuk file A dan B, kecuali skip_sheets
        def clean_name(sheet):
            return ''.join(re.findall(r'[A-Za-z ]+', sheet)).strip()


        map_a = {clean_name(s): s for s in xls_a.sheet_names if s not in skip_sheets}
        map_b = {clean_name(s): s for s in xls_b.sheet_names if s not in skip_sheets}

        print(f"[DEBUG] Map A sheets (clean name -> original): {map_a}")
        print(f"[DEBUG] Map B sheets (clean name -> original): {map_b}")

        common_clean_names = set(map_a.keys()).intersection(map_b.keys())
        print(f"[DEBUG] Common clean sheet names: {common_clean_names}")

        for clean_sheet in common_clean_names:
            sheet_a = map_a[clean_sheet]
            sheet_b = map_b[clean_sheet]
            print(f"[DEBUG] Membandingkan sheet A '{sheet_a}' dengan sheet B '{sheet_b}'")
            df_a = load_sheet_with_header_offset(xls_a, sheet_a, header_row=8)
            df_b = load_sheet_with_header_offset(xls_b, sheet_b, header_row=8)
            results[clean_sheet] = compare_sheets_fast(df_a, df_b)

    else:
        raise ValueError("Mode perbandingan tidak dikenali.")

    print("[DEBUG] Selesai compare_excels")
    return results


# ==================== FITUR 2: FILTER KATALOG ====================
def filter_excel_by_criteria(file_path, referensi=None, kode_referensi=None, kategori=None, harga_min=None, harga_max=None, tahun_filter=None):
    wb = load_workbook(file_path)
    summary = {}

    for sheet in wb.sheetnames[3:]:  # mulai sheet ke-4
        ws = wb[sheet]

        data = list(ws.values)
        columns = data[8]  # header di baris 9
        rows = data[9:]    # data mulai baris 10

        df = pd.DataFrame(rows, columns=columns)
        df.columns = df.columns.str.strip()

        # Filter Referensi (bisa banyak)
        if referensi and 'Referensi' in df.columns:
            df = df[df['Referensi'].isin(referensi)]

        # Filter Kode Referensi (bisa banyak)
        if kode_referensi and 'Kode Referensi' in df.columns:
            df = df[df['Kode Referensi'].isin(kode_referensi)]

        # Filter Kategori
        if kategori and 'Kategori' in df.columns:
            df = df[df['Kategori'] == kategori]

        # Filter harga
        if harga_min is not None and 'Harga' in df.columns:
            df = df[df['Harga'] >= harga_min]
        if harga_max is not None and 'Harga' in df.columns:
            df = df[df['Harga'] <= harga_max]

        # Filter tahun
        if tahun_filter and 'Tahun' in df.columns:
            df = df[df['Tahun'].isin(tahun_filter)]

        summary[sheet] = len(df)

        # Hapus data lama
        for row in ws.iter_rows(min_row=10, max_row=ws.max_row):
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.value = None

        # Tulis data hasil filter
        for r_idx, row in enumerate(df.itertuples(index=False), start=10):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    return wb, summary


# ==================== ROUTES ====================
@app.route('/')
def home():
    return render_template('index.html')

from flask import Flask, request, render_template, session
import os, uuid, pickle

@app.route('/compare', methods=['GET', 'POST'])
def compare():
    if request.method == 'POST':
        file_a = request.files.get('fileA')
        file_b = request.files.get('fileB')
        mode = request.form.get('mode')

        if not file_a or not file_b:
            return "File A dan File B wajib diunggah.", 400

        # Load ExcelFile objek untuk debug dan ambil sheet ke-4 (index 3)
        xls_a = pd.ExcelFile(file_a)
        xls_b = pd.ExcelFile(file_b)

        sheet_a_name = xls_a.sheet_names[3]  # sheet ke-4 file A
        sheet_b_name = xls_b.sheet_names[3]  # sheet ke-4 file B

        print("File A sheets:", xls_a.sheet_names)
        print("File B sheets:", xls_b.sheet_names)
        print("Sheet A yang dipilih (ke-4):", sheet_a_name)
        print("Sheet B yang dipilih (ke-4):", sheet_b_name)

        results = compare_excels(file_a, file_b, mode, sheet_a_name, sheet_b_name)

        summary = {sheet: len(df) for sheet, df in results.items()}

        session_id = str(uuid.uuid4())
        os.makedirs("tmp", exist_ok=True)
        with open(f"tmp/{session_id}.pkl", "wb") as f:
            pickle.dump(results, f)
        session["comparison_file"] = session_id

        return render_template('result.html', summary=summary)

    return render_template('compare.html')

@app.route('/filter', methods=['GET', 'POST'])
def filter():
    if request.method == 'POST':
        file = request.files['file']
        filename = file.filename
        os.makedirs("tmp", exist_ok=True)
        save_path = os.path.join("tmp", filename)
        file.save(save_path)

        # Ambil input filter
        referensi_raw = request.form.get('referensi', '').strip()
        kode_referensi_raw = request.form.get('kode_referensi', '').strip()
        kategori = request.form.get('kategori', '').strip()
        harga_min_raw = request.form.get('harga_min', '').strip()
        harga_max_raw = request.form.get('harga_max', '').strip()
        tahun_raw = request.form.get('tahun_filter', '').strip()

        # Parsing referensi jadi list
        referensi = [r.strip() for r in referensi_raw.split(',') if r.strip()] if referensi_raw else None
        kode_referensi = [k.strip() for k in kode_referensi_raw.split(',') if k.strip()] if kode_referensi_raw else None

        # Parsing harga
        try:
            harga_min = float(harga_min_raw) if harga_min_raw else None
            harga_max = float(harga_max_raw) if harga_max_raw else None
        except ValueError:
            return "Input harga tidak valid."

        # Parsing tahun
        try:
            tahun_filter = [int(t.strip()) for t in tahun_raw.split(',') if t.strip().isdigit()] if tahun_raw else None
        except ValueError:
            return "Input tahun tidak valid."

        # Jalankan filter
        wb, summary = filter_excel_by_criteria(
            save_path,
            referensi=referensi,
            kode_referensi=kode_referensi,
            kategori=kategori,
            harga_min=harga_min,
            harga_max=harga_max,
            tahun_filter=tahun_filter
        )

        # Simpan hasil
        session_id = str(uuid.uuid4())
        filtered_file_path = f"tmp/{session_id}_filtered.xlsx"
        wb.save(filtered_file_path)

        session["filtered_file_path"] = filtered_file_path
        session["filter_summary"] = summary

        return render_template('filter_result.html', summary=summary)

    return render_template('filter.html')


@app.route('/export')
def export():
    session_id = session.get("comparison_file")
    if not session_id:
        return "Tidak ada data untuk diekspor."

    pkl_path = f"tmp/{session_id}.pkl"
    if not os.path.exists(pkl_path):
        return "Tidak ada data untuk diekspor."

    with open(pkl_path, "rb") as f:
        results = pickle.load(f)

    export_folder = "tmp"
    os.makedirs(export_folder, exist_ok=True)
    filename = os.path.join(export_folder, f"hasil_perbandingan_export_{session_id}.xlsx")

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        sheet_written = False
        for sheet_name, data in results.items():
            df = pd.DataFrame(data)

            # Filter hanya data yang ditemukan
            if "Keterangan" in df.columns:
                df = df[df["Keterangan"] != "Tidak ditemukan"]

            # Jika df kosong, jangan tulis
            if not df.empty:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                sheet_written = True

        # Jika tidak ada sheet yang ditulis, buat sheet dummy supaya tidak error
        if not sheet_written:
            pd.DataFrame({"Info": ["Tidak ada data ditemukan untuk diekspor"]}).to_excel(writer, sheet_name="Info", index=False)

    return send_file(filename, as_attachment=True)


@app.route('/export_filter')
def export_filter():
    filtered_file_path = session.get("filtered_file_path")
    if not filtered_file_path or not os.path.exists(filtered_file_path):
        return "File hasil filter tidak ditemukan. Silakan lakukan filter ulang."

    return send_file(
        filtered_file_path,
        as_attachment=True,
        download_name="filtered_result.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
if __name__ == '__main__':
    app.run(debug=True)
